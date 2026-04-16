"""
Trustpilot Scraper for SBB using Selenium.
Extracts each review by clicking that review's own "See more" control and reading the modal text.
"""

from pathlib import Path
import logging
import os
import random
import re
import shutil
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment


logger = logging.getLogger(__name__)


def _close_modal(driver):
    """Close the currently open Trustpilot review modal if present.

    Args:
        driver: Selenium WebDriver instance.

    Returns:
        None.
    """
    try:
        close_button = driver.find_element(
            By.CSS_SELECTOR, "button[aria-label='Close']")
        driver.execute_script("arguments[0].click();", close_button)
        time.sleep(0.2)
    except (NoSuchElementException, StaleElementReferenceException):
        logger.debug("Modal close button not available; continuing")
    except WebDriverException as exc:
        logger.debug("Failed to close modal cleanly: %s", exc)


def _extract_review_text_from_modal(driver):
    """Extract review text from an open Trustpilot modal dialog.

    The function prefers the longest paragraph inside the modal and falls back
    to the modal's full inner text when needed.

    Args:
        driver: Selenium WebDriver instance.

    Returns:
        str: Cleaned review text, or an empty string if extraction fails.
    """
    try:
        modal = driver.find_element(By.XPATH, "//div[@role='dialog']")
    except (NoSuchElementException, WebDriverException):
        return ""

    best_text = ""
    try:
        for paragraph in modal.find_elements(By.TAG_NAME, "p"):
            try:
                text = driver.execute_script(
                    "return arguments[0].innerText;", paragraph)
                text = text.strip() if text else ""
                # Heuristic: modal's longest paragraph is usually the full review body.
                if len(text) > len(best_text):
                    best_text = text
            except WebDriverException:
                continue
    except WebDriverException as exc:
        logger.debug("Failed while scanning modal paragraphs: %s", exc)

    if best_text:
        return best_text.replace("See more", "").replace("... See more", "").strip()

    try:
        text = driver.execute_script("return arguments[0].innerText;", modal)
        return text.replace("See more", "").replace("... See more", "").strip() if text else ""
    except WebDriverException:
        return ""


def _extract_review_from_article(driver, article):
    """Extract structured review data from one Trustpilot article element.

    Args:
        driver: Selenium WebDriver instance.
        article: Selenium WebElement representing one Trustpilot review card.

    Returns:
        dict: Review fields with keys: author, title, rating, date, review.
    """
    review_data = {
        "author": "Unknown",
        "title": "",
        "rating": 0,
        "date": "",
        "review": "",
    }

    try:
        author_elem = article.find_element(
            By.CSS_SELECTOR, "span[data-consumer-name='true']")
        review_data["author"] = author_elem.text.strip() or "Unknown"
    except (NoSuchElementException, StaleElementReferenceException):
        try:
            aside = article.find_element(By.CSS_SELECTOR, "aside")
            aria_label = aside.get_attribute("aria-label") or ""
            if "Info for" in aria_label:
                review_data["author"] = aria_label.replace(
                    "Info for", "").strip()
        except (NoSuchElementException, StaleElementReferenceException):
            logger.debug("Author element not found in article")

    try:
        title_elem = article.find_element(By.TAG_NAME, "h2")
        review_data["title"] = title_elem.text.strip()
    except (NoSuchElementException, StaleElementReferenceException):
        logger.debug("Title element not found in article")

    try:
        rating_img = article.find_element(By.CSS_SELECTOR, "img[alt*='Rated']")
        alt = rating_img.get_attribute("alt") or ""
        # Parse rating with strict pattern first, then a lenient fallback.
        match = re.search(r"Rated\s+(\d)\s+out\s+of\s+5",
                          alt, flags=re.IGNORECASE)
        if not match:
            match = re.search(r"(\d)", alt)
        if match:
            review_data["rating"] = int(match.group(1))
    except (NoSuchElementException, StaleElementReferenceException, ValueError):
        logger.debug("Unable to parse rating from article")

    try:
        time_elem = article.find_element(By.TAG_NAME, "time")
        date = time_elem.get_attribute("datetime") or ""
        review_data["date"] = date.split("T")[0] if date else ""
    except (NoSuchElementException, StaleElementReferenceException):
        logger.debug("Date element not found in article")

    review_text = ""
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", article)
        time.sleep(0.2)

        see_more_candidates = article.find_elements(
            By.XPATH,
            ".//*[self::button or self::span or self::a][contains(normalize-space(.), 'See more')]",
        )
        if see_more_candidates:
            # Prefer modal text because card preview may be truncated.
            driver.execute_script(
                "arguments[0].click();", see_more_candidates[0])
            time.sleep(1)
            review_text = _extract_review_text_from_modal(driver)
            _close_modal(driver)
            time.sleep(0.2)

        if not review_text:
            # Fallback to card text when no modal text is available.
            preferred_paragraphs = article.find_elements(
                By.CSS_SELECTOR, "p[data-relevant-review-text-typography='true']"
            )
            if not preferred_paragraphs:
                preferred_paragraphs = article.find_elements(By.TAG_NAME, "p")

            best_text = ""
            for paragraph in preferred_paragraphs:
                try:
                    text = driver.execute_script(
                        "return arguments[0].innerText;", paragraph) or ""
                    text = text.strip()
                    if len(text) > len(best_text):
                        best_text = text
                except (StaleElementReferenceException, WebDriverException):
                    continue
            review_text = best_text
    except (StaleElementReferenceException, WebDriverException) as exc:
        logger.debug("Review text extraction failed for one article: %s", exc)

    review_data["review"] = review_text.replace(
        "See more", "").replace("... See more", "").strip()
    return review_data


def _resolve_chromedriver_path(explicit_path=None):
    """Resolve a valid ChromeDriver executable path.

    Resolution priority:
    1) explicit_path argument
    2) CHROMEDRIVER_PATH environment variable
    3) chromedriver from PATH
    4) /usr/bin/chromedriver

    Args:
        explicit_path: Optional explicit path to chromedriver.

    Returns:
        str: Resolved path to a chromedriver executable file.

    Raises:
        FileNotFoundError: If no valid chromedriver path is found.
    """
    candidates = [
        explicit_path,
        os.environ.get("CHROMEDRIVER_PATH"),
        # Prefer container-managed chromedriver on PATH for reproducibility.
        shutil.which("chromedriver"),
        "/usr/bin/chromedriver",
    ]

    for candidate in candidates:
        if not candidate:
            continue
        candidate_path = Path(candidate)
        if candidate_path.is_file():
            return str(candidate_path)

    raise FileNotFoundError(
        "No chromedriver found. In this devcontainer, install chromium-driver or set CHROMEDRIVER_PATH."
    )


def scrape_trustpilot_reviews(target_url, max_pages=2, max_reviews=50, chromedriver_path=None):
    """Scrape Trustpilot reviews from a company page.

    Args:
        target_url: Base Trustpilot company review URL.
        max_pages: Maximum number of paginated review pages to process.
        max_reviews: Maximum number of unique reviews to return.
        chromedriver_path: Optional explicit path to chromedriver.

    Returns:
        list[dict]: List of extracted review dictionaries.

    Raises:
        ValueError: If max_pages or max_reviews are not positive integers.
    """
    # print("=" * 70)
    # print("Trustpilot Scraper (per-review modal extraction)")
    # print("=" * 70)

    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    # Devcontainer runs without a display server, so headless mode is required.
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument(
        "--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option(
        "excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    chrome_options.add_argument("--disable-images")

    reviews_list = []
    seen_keys = set()
    driver = None

    if max_pages <= 0:
        raise ValueError("max_pages must be greater than 0")
    if max_reviews <= 0:
        raise ValueError("max_reviews must be greater than 0")

    try:
        chromedriver_path = _resolve_chromedriver_path(chromedriver_path)
        print(f"   Using chromedriver: {chromedriver_path}")
        service = Service(chromedriver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print("   Browser initialized\n")

        for page_num in range(1, max_pages + 1):
            if len(reviews_list) >= max_reviews:
                break

            url_page = f"{target_url}?page={page_num}"
            print(f"Page {page_num}: {url_page}")
            try:
                driver.get(url_page)
            except (TimeoutException, WebDriverException) as exc:
                logger.warning(
                    "Skipping page %s due to navigation error: %s", page_num, exc)
                continue
            # Small jitter helps lazy-rendered sections settle before DOM queries.
            time.sleep(random.uniform(3, 5))

            for _ in range(3):
                # Multiple viewport passes trigger lazy loading for additional cards.
                driver.execute_script(
                    "window.scrollBy(0, window.innerHeight);")
                time.sleep(0.25)

            articles = driver.find_elements(By.TAG_NAME, "article")
            print(f"   Found {len(articles)} reviews on page {page_num}")

            for article in articles:
                if len(reviews_list) >= max_reviews:
                    break

                try:
                    review_data = _extract_review_from_article(driver, article)
                    review_key = (
                        review_data["author"],
                        review_data["title"],
                        review_data["date"],
                        review_data["review"],
                    )
                    # De-duplicate overlapping cards across pages/reloads.
                    if review_key in seen_keys:
                        continue
                    # Keep entries with rating even if body text is short or empty.
                    if review_data["review"] or review_data["rating"] > 0:
                        seen_keys.add(review_key)
                        reviews_list.append({
                            "id": len(reviews_list) + 1,
                            **review_data,
                        })
                        print(
                            f"   [{len(reviews_list)}] {review_data['author']} - rating={review_data['rating']} - {review_data['review'][:40]}"
                        )
                except (StaleElementReferenceException, WebDriverException) as exc:
                    logger.debug(
                        "Skipping one article due to extraction error: %s", exc)
                    continue

    except (WebDriverException, FileNotFoundError) as exc:
        logger.exception("Scraper failed to initialize or run: %s", exc)
    except Exception as exc:
        logger.exception("Unexpected scraper failure: %s", exc)
    finally:
        if driver:
            driver.quit()
        print("\nBrowser closed")

    return reviews_list


def save_excel(reviews, output_file="data/trustpilot_reviews.xlsx"):
    """Write scraped reviews to an Excel file.

    Args:
        reviews: List of review dictionaries produced by the scraper.
        output_file: Destination .xlsx file path.

    Returns:
        None.
    """
    print(f"\nSaving {len(reviews)} reviews to Excel...\n")

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Reviews"

    # Header labels are business-facing names, while internal keys stay stable in review dicts.
    headers = ["ID", "Author", "Title", "Rating", "Date", "Review"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row, review in enumerate(reviews, 2):
        ws.cell(row=row, column=1, value=review["id"])
        ws.cell(row=row, column=2, value=review["author"])
        ws.cell(row=row, column=3, value=review["title"])
        ws.cell(row=row, column=4, value=review["rating"])
        ws.cell(row=row, column=5, value=review["date"])
        ws.cell(row=row, column=6, value=review["review"])
        ws.cell(row=row, column=6).alignment = Alignment(
            wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 65

    try:
        wb.save(output_path)
    except OSError as exc:
        logger.exception(
            "Failed to save Excel output at %s: %s", output_path, exc)
        raise
    print(f"Saved to: {output_path}\n")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )

    TARGET_URL = "https://www.trustpilot.com/review/www.sbb.ch"
    reviews = scrape_trustpilot_reviews(
        TARGET_URL, max_pages=11, max_reviews=200)

    print("=" * 70)
    if reviews:
        print(f"SUCCESS: Scraped {len(reviews)} reviews")
        save_excel(reviews)
    else:
        print("FAILED: No reviews scraped")
    print("=" * 70)
